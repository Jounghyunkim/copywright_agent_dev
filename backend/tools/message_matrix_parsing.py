"""Message Matrix Excel 파일을 파싱하여 마크다운으로 변환하는 모듈."""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ── 데이터 구조 ──────────────────────────────────────────────

@dataclass
class USP:
    usp_no: str
    feature_name: str
    key_message_full: str
    key_message_short: str
    benefit_description: str
    rtb: str
    disclaimer: str
    certification: str
    remark: str


@dataclass
class BenefitCategory:
    number: str
    name: str
    key_message: str
    usps: list[USP] = field(default_factory=list)


@dataclass
class ProductInfo:
    product_name: str
    sub_name: str
    head_message: str
    description: str
    categories: list[BenefitCategory] = field(default_factory=list)


# ── 셀 값 안전하게 읽기 ──────────────────────────────────────

def _cell(ws: Worksheet, coord: str) -> str:
    """셀 값을 문자열로 반환. 수식(=LEN 등)이나 None은 빈 문자열로."""
    val = ws[coord].value
    if val is None:
        return ""
    val = str(val)
    if val.startswith("="):
        return ""
    return val.strip()


def _cell_rc(ws: Worksheet, row: int, col: int) -> str:
    """(row, col) 기반으로 셀 값을 문자열로 반환."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    val = str(val)
    if val.startswith("="):
        return ""
    return val.strip()


# ── Excel 파싱 ───────────────────────────────────────────────

# 고정 컬럼 매핑 (1-indexed)
COL_CATEGORY_NO = 2       # B
COL_CATEGORY_NAME = 3     # C
COL_CATEGORY_MSG = 4      # D
COL_USP_NO = 5            # E
COL_FEATURE_NAME = 6      # F
COL_KEY_MSG_FULL = 10     # J
COL_KEY_MSG_SHORT = 14    # N
COL_BENEFIT_DESC = 18     # R
COL_RTB = 22              # V
COL_DISCLAIMER = 23       # W
COL_CERTIFICATION = 24    # X
COL_REMARK = 25           # Y

DATA_START_ROW = 14


def _parse_sheet(ws: Worksheet) -> ProductInfo:
    """단일 시트를 파싱하여 ProductInfo를 반환."""
    product = ProductInfo(
        product_name=_cell(ws, "B3"),
        sub_name=_cell(ws, "B4"),
        head_message=_cell(ws, "J7"),
        description=_cell(ws, "J9"),
    )

    current_cat: BenefitCategory | None = None

    for row in range(DATA_START_ROW, ws.max_row + 1):
        cat_no = _cell_rc(ws, row, COL_CATEGORY_NO)
        cat_name = _cell_rc(ws, row, COL_CATEGORY_NAME)
        usp_no = _cell_rc(ws, row, COL_USP_NO)
        feature = _cell_rc(ws, row, COL_FEATURE_NAME)

        # 새 카테고리 시작
        if cat_no and cat_name:
            current_cat = BenefitCategory(
                number=cat_no,
                name=cat_name,
                key_message=_cell_rc(ws, row, COL_CATEGORY_MSG),
            )
            product.categories.append(current_cat)

        # USP 행
        if usp_no and feature:
            usp = USP(
                usp_no=usp_no,
                feature_name=feature,
                key_message_full=_cell_rc(ws, row, COL_KEY_MSG_FULL),
                key_message_short=_cell_rc(ws, row, COL_KEY_MSG_SHORT),
                benefit_description=_cell_rc(ws, row, COL_BENEFIT_DESC),
                rtb=_cell_rc(ws, row, COL_RTB),
                disclaimer=_cell_rc(ws, row, COL_DISCLAIMER),
                certification=_cell_rc(ws, row, COL_CERTIFICATION),
                remark=_cell_rc(ws, row, COL_REMARK),
            )
            if current_cat is None:
                # 카테고리 없이 USP만 있는 경우 — USP 번호 앞자리로 소속 카테고리 탐색
                if product.categories:
                    cat_prefix = usp_no.split("-")[0]
                    matched = [c for c in product.categories if c.number == cat_prefix]
                    if matched:
                        matched[-1].usps.append(usp)
                    else:
                        product.categories[-1].usps.append(usp)
                else:
                    current_cat = BenefitCategory(
                        number="?", name="Unknown", key_message=""
                    )
                    current_cat.usps.append(usp)
                    product.categories.append(current_cat)
            else:
                current_cat.usps.append(usp)

        # 빈 행이면 current_cat 리셋
        if not cat_no and not cat_name and not usp_no:
            current_cat = None

    return product


def get_sheet_names(filepath: str | Path) -> list[str]:
    """Excel 파일의 시트 이름 목록을 반환."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def parse_excel(
    filepath: str | Path,
    sheet_names: list[str] | None = None,
) -> dict[str, ProductInfo]:
    """Message Matrix Excel 파일을 파싱하여 {시트명: ProductInfo} 딕셔너리를 반환.

    Args:
        filepath: 입력 Excel 파일 경로.
        sheet_names: 파싱할 시트 이름 목록. None이면 모든 시트를 파싱.
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    targets = sheet_names if sheet_names else wb.sheetnames
    results: dict[str, ProductInfo] = {}
    for name in targets:
        ws: Worksheet = wb[name]  # type: ignore[assignment]
        results[name] = _parse_sheet(ws)
    wb.close()
    return results


# ── 마크다운 생성 ────────────────────────────────────────────

def _escape_md_table(text: str) -> str:
    """테이블 셀 내 줄바꿈과 파이프 이스케이프."""
    return text.replace("|", "\\|").replace("\n", " ").strip()


def _summarize_rtb(rtb: str) -> list[str]:
    """RTB 텍스트를 불릿 포인트로 분리."""
    if not rtb:
        return []
    lines = []
    for line in rtb.split("\n"):
        line = line.strip()
        if not line:
            continue
        if line.startswith("[") and "]" in line:
            lines.append(line)
        elif line.startswith("-") or line.startswith("*"):
            lines.append(line.lstrip("-* "))
        else:
            lines.append(line)
    return lines


def generate_markdown(product: ProductInfo) -> str:
    """ProductInfo를 템플릿 형식의 마크다운 문자열로 변환."""
    lines: list[str] = []
    w = lines.append

    full_name = product.product_name
    if product.sub_name:
        full_name += f" {product.sub_name}"

    w("# Message Matrix Parsing 분석")
    w("")
    w("## 제품 정보")
    w("")
    w("| 항목 | 내용 |")
    w("|------|------|")
    w(f"| **제품명** | {full_name} |")
    w(f"| **Product Head Message** | {product.head_message} |")
    w(f"| **Product Description** | {product.description} |")
    w("")
    w("---")
    w("")

    # 매트릭스 구조 (고정)
    w("## 매트릭스 구조")
    w("")
    w("### 컬럼 구성")
    w("")
    w("| 컬럼 | 설명 | 용도 / 제약 |")
    w("|------|------|-------------|")
    w("| BENEFIT CATEGORY | 상위 카테고리 | 기능 그룹핑 |")
    w("| CATEGORY KEY MESSAGE | 카테고리별 핵심 메시지 | 카테고리 요약 |")
    w("| USP # | 하위 기능 번호 | 기능 식별자 |")
    w("| USP / FEATURE NAME | 기능명 | Max 27 Bytes |")
    w("| KEY MESSAGE (Full) | 풀 카피 | Max 60 Bytes, LG.com Headline 용 |")
    w("| KEY MESSAGE (Short) | 숏 카피 | Max 60 Bytes, TVC/USP Film/Event/POP/Social 용 |")
    w("| BENEFIT DESCRIPTION | 한 문장 설명 | Max 100 Bytes, LG.com Body Copy 용 |")
    w("| REASON TO BELIEVE | 기술적 근거 상세 설명 | 제품 소개/교육 자료 용 |")
    w("| DISCLAIMER | 법적 고지/조건 | 제한 없음 |")
    w("| CERTIFICATION | 인증 정보 | - |")
    w("| REMARK | 비고 | - |")
    w("")
    w("---")
    w("")

    # 카테고리별 USP 상세
    w("## Benefit Category 별 USP 상세")
    w("")

    total_usps = 0
    for cat in product.categories:
        total_usps += len(cat.usps)
        w(f"### Category {cat.number}: {cat.name}")
        w("")
        if cat.key_message:
            w(f"> {_escape_md_table(cat.key_message)}")
            w("")

        # USP 테이블
        w("| USP # | Feature Name | Key Message (Full) | Key Message (Short) | Benefit Description |")
        w("|-------|-------------|-------------------|---------------------|---------------------|")
        for usp in cat.usps:
            w(
                f"| {usp.usp_no} "
                f"| {_escape_md_table(usp.feature_name)} "
                f"| {_escape_md_table(usp.key_message_full)} "
                f"| {_escape_md_table(usp.key_message_short)} "
                f"| {_escape_md_table(usp.benefit_description)} |"
            )
        w("")

        # RTB 상세
        has_rtb = any(usp.rtb for usp in cat.usps)
        if has_rtb:
            w("#### RTB (Reason to Believe) 상세")
            w("")
            for usp in cat.usps:
                if not usp.rtb:
                    continue
                w(f"- **{_escape_md_table(usp.feature_name)}**")
                for bullet in _summarize_rtb(usp.rtb):
                    w(f"  - {bullet}")
                w("")

        w("---")
        w("")

    # 데이터 특성 요약
    w("## 데이터 특성 요약")
    w("")
    w("| 항목 | 값 |")
    w("|------|-----|")
    w(f"| Benefit Category 수 | {len(product.categories)} |")
    w(f"| 총 USP/Feature 수 | {total_usps} |")
    w("| 바이트 제한 관리 | LEN / LENB 수식으로 CHAR / BYTES 자동 계산 |")
    w("| 카피 변형 | Full (LG.com) / Short (TVC·소셜) / Benefit Description (본문) |")
    w("| 언어 | 영문 기반 (일부 한국어 혼용) |")
    w("")

    # 카피라이팅 에이전트 활용 시사점
    w("## 카피라이팅 에이전트 활용 시사점")
    w("")
    w("1. **입력 구조**: Brief에서 제품명·카테고리·USP 계층 구조를 반영해야 함")
    w("2. **바이트 제한**: 각 카피 유형별 바이트 제한(27/60/100)이 존재하므로, 생성 시 바이트 수 검증 필요")
    w("3. **카피 변형 계층**: 동일 USP에 대해 Full → Short → Benefit Description → RTB 순으로 상세도가 증가하는 계층 구조")
    w("4. **RTB-Disclaimer 쌍**: 모든 기능 주장에는 RTB(근거)와 Disclaimer(법적 고지)가 쌍으로 존재")
    w("5. **다채널 용도**: 카피가 LG.com, TVC, 소셜, POP 등 채널별로 분화되어 사용됨")
    w("")

    return "\n".join(lines)


# ── CLI 진입점 ───────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Message Matrix Excel → Markdown 변환"
    )
    parser.add_argument(
        "input",
        help="입력 Excel 파일 경로 (.xlsx)",
    )
    parser.add_argument(
        "-o", "--output-dir",
        help="출력 디렉토리 경로 (미지정 시 입력 파일과 같은 디렉토리)",
        default=None,
    )
    args = parser.parse_args()

    # 시트 목록 표시 및 사용자 선택
    all_sheets = get_sheet_names(args.input)

    if len(all_sheets) == 1:
        selected_sheets = all_sheets
        print(f"시트 1개 감지: {all_sheets[0]}")
    else:
        print(f"\n총 {len(all_sheets)}개 시트가 발견되었습니다:\n")
        for i, name in enumerate(all_sheets, 1):
            print(f"  {i}. {name}")
        print()
        user_input = input("변환할 시트 번호를 입력하세요 (예: 1,3,5 / 전체: all): ").strip()

        if user_input.lower() == "all":
            selected_sheets = all_sheets
        else:
            indices = [int(x.strip()) for x in user_input.split(",") if x.strip().isdigit()]
            selected_sheets = [
                all_sheets[i - 1] for i in indices if 1 <= i <= len(all_sheets)
            ]
            if not selected_sheets:
                print("유효한 시트 번호가 없습니다.")
                return

        print(f"\n선택된 시트: {', '.join(selected_sheets)}\n")

    # 파싱 및 마크다운 생성
    products = parse_excel(args.input, sheet_names=selected_sheets)
    output_dir = Path(args.output_dir) if args.output_dir else Path(args.input).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name, product in products.items():
        md = generate_markdown(product)
        safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        out_path = output_dir / f"{safe_name}.md"
        out_path.write_text(md, encoding="utf-8")
        print(f"생성 완료: {out_path}")


if __name__ == "__main__":
    main()
